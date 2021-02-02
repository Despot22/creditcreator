import openpyxl, os, re, docx

#set current working directory
os.chdir('C:\\02 Python\\Credit Creation')

#load excel workbook
wb = openpyxl.load_workbook('testdoc.xlsx')

#select first worksheet
sheet = wb.active

#import relevant columns (starting with first 30)
creditList = []


#for i in sheet.iter_rows():
for i in range(2, sheet.max_row):
        #find credit
        credit = sheet.cell(row=i, column=8).value
        #find page
        page = sheet.cell(row=i, column=29).value
        fullCredit = credit + ' ' + page
        fullCredit.strip()
        creditList.append(fullCredit)
 

#reformat alphabetically
sortedList = sorted(creditList)
finalList = ", ".join(sortedList)



#export data
d = docx.Document()
d.add_paragraph(finalList)
d.save('C:\\02 Python\\Credit Creation\\credits.docx')
